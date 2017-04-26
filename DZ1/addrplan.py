import glob
import ipaddress
import re
from openpyxl import Workbook

Mega_set=set()  #итоговое множество кортежей вида (Сеть/Маска, Шлюз, Описание)
IP_set=set()  #множество кортежей (Сеть/Маска, IPaddr), полученных из команды 'ip address'
GW_set=set()  #множество шлюзов (GW_Addr)
Net_set=set()  #множество сетей, для которых явно заданы шлюзы
def Classify (stroka):
    IP=re.match(' ip address (([0-9]{1,3}\.){3}[0-9]{1,3}) (([0-9]{1,3}\.){3}[0-9]{1,3})',stroka)
    if IP:
        IP_set.add((ipaddress.IPv4Network((IP.group(1), IP.group(3)), strict=False), IP.group(1)))
        return
    GW=re.match('ip default-gateway (.*)',stroka)
    if GW:
        GW_set.add(ipaddress.IPv4Address(GW.group(1)))
        return
    NH=re.match('ip route 0.0.0.0 0.0.0.0 (.*)',stroka)
    if NH:
        GW_set.add(ipaddress.IPv4Address(NH.group(1)))
        return
    VRRP=re.match(' vrrp (.+) ip (.*)',stroka)
    if VRRP:
        GW_set.add(ipaddress.IPv4Address(VRRP.group(2)))
        return
    GLBP = re.match(' glbp (.+) ip (.*)', stroka)
    if GLBP:
        GW_set.add(ipaddress.IPv4Address(GLBP.group(2)))
        return
    HSRP = re.match(' standby (.+) ip (.*)', stroka)
    if HSRP:
        GW_set.add(ipaddress.IPv4Address(HSRP.group(2)))
        return

list_of_files=glob.glob('C:\\work\\Python\\p4ne_training\\config_files\\*.txt')  #составляет список файлов с расширением .txt
for i in list_of_files:  #пробегаем по списку файлов
    with open (i) as f:  #открываем файл
        lines=f.readlines()  #формируем список строк файла
        for j in lines:  #пробегаем по строкам
            Classify(j)

for gateway in GW_set:
   for net_mask_gw in IP_set:
       if gateway in net_mask_gw[0]:
        Mega_set.add((net_mask_gw[0],gateway,'Explicit gateway configuration'))  #формируем итоговое множество кортежей
        Net_set.add(net_mask_gw[0])  #формируем множество сетей, для которых известен шлюз
for net_mask_gw in IP_set:
    if not(net_mask_gw[0] in Net_set):
        Mega_set.add((net_mask_gw[0],net_mask_gw[1],'Inter-router Link or stub network'))
def Sortfunc(ipnet):  #функция для сортировки списка
    return int(ipnet[0].network_address._ip)
Megalist=list(Mega_set)
#print('Network/Mask, Gateway, Description')
#for i in sorted(Megalist,key=Sortfunc):
#    print(i)

wb=Workbook()
ws=wb.active
ws.title='Address Plan'
ws['A1']='Network/Mask'
ws['B1']='Gateway'
ws['C1']='Description'
for i in sorted(Megalist, key=Sortfunc):
    ws.append([str(i[0]), str(i[1]), str(i[2])])
wb.save('C:\\work\\Python\\Addrplan.xlsx')
